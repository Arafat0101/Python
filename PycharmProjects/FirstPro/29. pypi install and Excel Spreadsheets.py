#openpyxl install
#Website: https://pypi.org/project/openpyxl/
# open terminal and write command: pip install openpyxl


# from openpyxl import Workbook
#
# workbook = Workbook()
# sheet = workbook.active
#
# sheet["A1"] = "hello"
# sheet["B1"] = "world!"
#
# workbook.save(filename="hello_world.xlsx")


import openpyxl as xl
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['A1']
#cell = sheet.cell(1, 1)
#print(cell.value)
print(sheet.max_row)

for row in range(1, sheet.max_row+1):
    #print(row)
    cell = sheet.cell(row,3)
    print(cell.value)